from __future__ import annotations

from dataclasses import dataclass
from itertools import cycle
from pathlib import Path

from openpyxl import load_workbook

from .models import AssetPlan, CreativeText, PlatformConfig, RuntimeSettings, TemplateMap


@dataclass(frozen=True)
class SpreadsheetCreative:
    row_number: int
    header_text: str
    subtitle_text: str
    image_path: Path
    image_id: str


def _image_anchor_row(image) -> int | None:
    anchor = getattr(image, "anchor", None)
    anchor_from = getattr(anchor, "_from", None)
    if anchor_from is None:
        return None
    return anchor_from.row + 1


def import_spreadsheet_creatives(
    workbook_path: Path,
    workspace: Path,
    template_map: TemplateMap,
) -> list[SpreadsheetCreative]:
    workbook = load_workbook(workbook_path)
    sheet = workbook[workbook.sheetnames[0]]
    headers = [sheet.cell(1, idx).value for idx in range(1, sheet.max_column + 1)]
    required = {"header_text", "subtitle_text", template_map.fields["background_image"]}
    missing = [field for field in required if field not in headers]
    if missing:
        raise RuntimeError(f"Spreadsheet is missing required columns: {', '.join(sorted(missing))}")

    images_by_row: dict[int, Path] = {}
    extracted_dir = workspace / "assets" / "images" / "xlsx_imports"
    extracted_dir.mkdir(parents=True, exist_ok=True)
    workbook_stem = workbook_path.stem
    for image in getattr(sheet, "_images", []):
        row_number = _image_anchor_row(image)
        if row_number is None:
            continue
        ext = image.format.lower()
        filename = f"{workbook_stem}_row_{row_number:03d}.{ext}"
        output_path = extracted_dir / filename
        output_path.write_bytes(image.ref.getvalue())
        images_by_row[row_number] = output_path

    index = {name: idx + 1 for idx, name in enumerate(headers) if name}
    creatives: list[SpreadsheetCreative] = []
    for row_number in range(2, sheet.max_row + 1):
        header_text = sheet.cell(row_number, index["header_text"]).value
        subtitle_text = sheet.cell(row_number, index["subtitle_text"]).value
        image_path = images_by_row.get(row_number)
        if not header_text and not subtitle_text and image_path is None:
            continue
        if not header_text or not subtitle_text or image_path is None:
            raise RuntimeError(
                f"Spreadsheet row {row_number} must include header_text, subtitle_text, and an embedded image"
            )
        creatives.append(
            SpreadsheetCreative(
                row_number=row_number,
                header_text=str(header_text).strip(),
                subtitle_text=str(subtitle_text).strip(),
                image_path=image_path,
                image_id=image_path.name,
            )
        )

    if not creatives:
        raise RuntimeError("Spreadsheet did not contain any complete creative rows")
    return creatives


def build_plans_from_spreadsheet(
    settings: RuntimeSettings,
    platforms: list[PlatformConfig],
    creatives: list[SpreadsheetCreative],
) -> list[AssetPlan]:
    plans: list[AssetPlan] = []
    creative_cycle = cycle(creatives)
    for platform in platforms:
        for sequence in range(1, platform.daily_count + 1):
            creative = next(creative_cycle)
            asset_id = f"{platform.name}_{settings.run_id}_{sequence:03d}"
            plans.append(
                AssetPlan(
                    run_id=settings.run_id,
                    asset_id=asset_id,
                    platform=platform.name,
                    sequence=sequence,
                    image_id=creative.image_id,
                    image_path=creative.image_path,
                    hook=CreativeText(
                        id=f"xlsx_row_{creative.row_number:03d}_header",
                        text=creative.header_text,
                        category="spreadsheet",
                        enabled=True,
                    ),
                    subtitle=CreativeText(
                        id=f"xlsx_row_{creative.row_number:03d}_subtitle",
                        text=creative.subtitle_text,
                        category="spreadsheet",
                        enabled=True,
                    ),
                    caption=platform.caption_patterns[(sequence - 1) % len(platform.caption_patterns)]
                    if platform.caption_patterns
                    else "",
                )
            )
    return plans
