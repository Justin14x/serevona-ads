from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from PIL import Image

from src.models import PlatformConfig, RuntimeSettings, TemplateMap
from src.xlsx_inputs import SpreadsheetCreative, build_plans_from_spreadsheet, import_spreadsheet_creatives


class XlsxInputTests(unittest.TestCase):
    def test_import_spreadsheet_creatives_extracts_embedded_image(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            workspace = Path(tmp_dir)
            workbook_path = workspace / "fields.xlsx"
            image_path = workspace / "seed.png"

            Image.new("RGB", (20, 20), color=(255, 128, 0)).save(image_path)
            wb = Workbook()
            ws = wb.active
            ws["A1"] = "header_text"
            ws["B1"] = "subtitle_text"
            ws["C1"] = "background_image:image"
            ws["A2"] = "header"
            ws["B2"] = "subtitle"
            ws.add_image(XLImage(str(image_path)), "C2")
            wb.save(workbook_path)

            creatives = import_spreadsheet_creatives(
                workbook_path=workbook_path,
                workspace=workspace,
                template_map=TemplateMap(
                    template_id="EAHEFZ9SogE",
                    fields={
                        "background_image": "background_image:image",
                        "header_text": "header_text",
                        "subtitle_text": "subtitle_text",
                    },
                ),
            )

            self.assertEqual(len(creatives), 1)
            self.assertEqual(creatives[0].header_text, "header")
            self.assertTrue(creatives[0].image_path.exists())

    def test_build_plans_from_spreadsheet_cycles_rows_for_platform_targets(self) -> None:
        settings = RuntimeSettings(
            workspace=Path("/tmp/workspace"),
            config_dir=Path("/tmp/workspace/config"),
            assets_dir=Path("/tmp/workspace/assets"),
            outputs_dir=Path("/tmp/workspace/outputs"),
            state_dir=Path("/tmp/workspace/state"),
            run_id="2026-03-16_daily",
            canva_mode="mock",
            same_pair_cooldown_days=7,
            image_cooldown_days=3,
            dry_run_uploads=True,
            sample=False,
            force=False,
            xlsx_source=Path("/tmp/workspace/fields.xlsx"),
        )
        plans = build_plans_from_spreadsheet(
            settings=settings,
            platforms=[PlatformConfig(name="instagram", daily_count=2, caption_patterns=["a", "b"])],
            creatives=[
                SpreadsheetCreative(
                    row_number=2,
                    header_text="header",
                    subtitle_text="subtitle",
                    image_path=Path("/tmp/workspace/assets/images/one.png"),
                    image_id="one.png",
                )
            ],
        )

        self.assertEqual(len(plans), 2)
        self.assertEqual(plans[0].hook.text, "header")
        self.assertEqual(plans[1].subtitle.text, "subtitle")


if __name__ == "__main__":
    unittest.main()
